"""
Seed: carga emergencias, protocolos y pasos de navegación desde el Excel.
Uso: conda run -n tesis python scripts/seed.py
"""
import asyncio
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

EXCEL_PATH = ROOT / "data" / "Protocolos y pasos.xlsx"

import openpyxl
from sqlalchemy import delete
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.infrastructure.config import settings
from app.infrastructure.database.models import EmergenciaModel, PasoModel, ProtocoloModel


def _limpiar(valor) -> str | None:
    if valor is None:
        return None
    s = str(valor).strip()
    return None if s.upper() == "NULL" or s == "" else s


def _partir(raw: str | None) -> tuple[str | None, str | None]:
    """Divide 'A\\nB' en ('A', 'B'). Si es un solo valor retorna (valor, None)."""
    if not raw:
        return None, None
    partes = [p.strip() for p in raw.split("\n") if p.strip() and p.strip().upper() != "NULL"]
    primero = partes[0] if len(partes) > 0 else None
    segundo = partes[1] if len(partes) > 1 else None
    return primero, segundo


def cargar_excel(path: Path):
    wb = openpyxl.load_workbook(path)

    # ── EMERGENCIAS ──────────────────────────────────────────────────────────
    emergencias = []
    for row in wb["EMERGENCIAS"].iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        emergencias.append({
            "id_emergencia":          str(row[0]).strip(),
            "nombre_emergencia":      str(row[1]).strip(),
            "descripcion_emergencia": str(row[2]).strip() if row[2] else "",
            "grupo_edad":             str(row[3]).strip() if row[3] else "",
            "severidad":              str(row[4]).strip() if row[4] else "",
            "etiqueta":               str(row[5]).strip() if row[5] else "",
            "evaluacion_inicial":     str(row[6]).strip() if row[6] else "",
        })

    # ── PROTOCOLOS ───────────────────────────────────────────────────────────
    protocolos = []
    for row in wb["PROTOCOLOS"].iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        id_protocolo  = str(row[0]).strip()
        id_emergencia = _limpiar(row[1])
        id_em_valido  = id_emergencia if (id_emergencia and id_emergencia.startswith("E")) else None

        protocolos.append({
            "id_protocolo":  id_protocolo,
            "id_emergencia": id_em_valido,
            "numero":        int(row[2]) if row[2] else 0,
            "instruccion":   str(row[3]).strip() if row[3] else "",
            "observacion":   _limpiar(row[4]),
            "imagen":        f"{id_protocolo}.png",
        })

    # ── PASOS (navegación) ───────────────────────────────────────────────────
    ids_protocolos = {p["id_protocolo"] for p in protocolos}
    pasos = []
    for row in wb["PASOS"].iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        id_protocolo = str(row[0]).strip()

        sig_si,  sig_no   = _partir(_limpiar(row[2]))   # paso_siguiente
        anx_si,  anx_no   = _partir(_limpiar(row[3]))   # anexo

        # Filtrar referencias a IDs que no existen en protocolo
        sig_si  = sig_si  if sig_si  in ids_protocolos else None
        sig_no  = sig_no  if sig_no  in ids_protocolos else None
        anx_si  = anx_si  if anx_si  in ids_protocolos else None
        anx_no  = anx_no  if anx_no  in ids_protocolos else None

        pasos.append({
            "id_protocolo":    id_protocolo,
            "paso_siguiente":   sig_si,
            "paso_siguiente_no": sig_no,
            "anexo_si":         anx_si,
            "anexo_no":         anx_no,
        })

    return emergencias, protocolos, pasos


async def seed():
    engine = create_async_engine(settings.database_url)
    SessionLocal = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    emergencias, protocolos, pasos = cargar_excel(EXCEL_PATH)

    async with SessionLocal() as session:
        # Limpiar en orden FK
        await session.execute(delete(PasoModel))
        await session.execute(delete(ProtocoloModel))
        await session.execute(delete(EmergenciaModel))
        await session.commit()

        for e in emergencias:
            session.add(EmergenciaModel(**e))
        await session.commit()
        print(f"  {len(emergencias)} emergencias insertadas.")

        for p in protocolos:
            session.add(ProtocoloModel(**p))
        await session.commit()
        print(f"  {len(protocolos)} protocolos insertados.")

        for p in pasos:
            session.add(PasoModel(**p))
        await session.commit()
        print(f"  {len(pasos)} pasos de navegación insertados.")

    await engine.dispose()
    print("Seed completado exitosamente.")


asyncio.run(seed())
